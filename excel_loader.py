<<<<<<< HEAD
# excel_loader.py
from __future__ import annotations

from pathlib import Path
from datetime import date, datetime
import re
import unicodedata

import openpyxl


def _normalize_name(s: str) -> str:
    s = re.sub(r"\s+", " ", str(s).strip())
    if " " not in s:
        s = re.sub(r"(?<!^)(?=[A-Z])", " ", s)
    return s.strip()


def compute_checksum_from12(d12: str) -> int:
    s_odd = sum(int(d12[i]) for i in range(0, 12, 2))
    s_even = sum(int(d12[i]) for i in range(1, 12, 2))
    total = s_odd + 3 * s_even
    return (10 - (total % 10)) % 10


def _to_date(v) -> date | None:
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _norm_colname(x) -> str:
    return re.sub(r"\s+", "", str(x).strip().lower())


def _find_col_index(headers: list, wanted: list[str]) -> int | None:
    norm = [_norm_colname(h) for h in headers]
    wanted_norm = [_norm_colname(w) for w in wanted]
    for w in wanted_norm:
        if w in norm:
            return norm.index(w)
    return None


def _find_sheet(wb, preferred: list[str]) -> str | None:
    names = wb.sheetnames
    norm_map = {n.lower().strip(): n for n in names}
    for p in preferred:
        k = p.lower().strip()
        if k in norm_map:
            return norm_map[k]
    return None


def load_members(xlsx_path: Path) -> list[dict]:
    """
    Leest leden uit Excel en geeft terug:
    [{"code13": "...", "name": "...", "qr": "...|None"}, ...]
    Accepteert kolommen:
      - code: "Lidcode" of "Code" of "Nummer"
      - naam: "Naam" of "Name"
      - qr:   "QR" of "Qr" of "Link" (optioneel)
    """
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    sheet_name = _find_sheet(wb, ["Leden", "Members"])
    if not sheet_name:
        wb.close()
        raise FileNotFoundError("Blad 'Leden' ontbreekt in leden.xlsx")

    ws = wb[sheet_name]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_code = _find_col_index(headers, ["Lidcode", "Code", "Nummer"])
    idx_name = _find_col_index(headers, ["Naam", "Name"])
    idx_qr = _find_col_index(headers, ["QR", "Qr", "Link", "Url", "URL"])

    if idx_code is None or idx_name is None:
        wb.close()
        raise KeyError(
            f"Verkeerde kolomnamen in blad '{sheet_name}'. "
            f"Gevonden: {headers}. Verwacht minimaal Code/Lidcode + Naam."
        )

    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        nummer = row[idx_code] if idx_code < len(row) else None
        naam = row[idx_name] if idx_name < len(row) else None
        qr = row[idx_qr] if (idx_qr is not None and idx_qr < len(row)) else None

        if not nummer or not naam:
            continue

        digits = "".join(ch for ch in str(nummer) if ch.isdigit())
        if len(digits) < 12:
            continue

        d12 = digits[:12]
        code13 = d12 + str(compute_checksum_from12(d12))

        out.append(
            {
                "code13": code13,
                "name": _normalize_name(naam),
                "qr": (str(qr).strip() if qr else None),
            }
        )

    wb.close()

    out.sort(
        key=lambda d: unicodedata.normalize("NFKD", d["name"])
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )
    return out


def load_members_df(xlsx_path: Path | None = None):
    """
    Voor menu_lidkaarten.py:
    geeft een DataFrame terug met kolommen CODE_COL en NAME_COL (uit config.py),
    ongeacht of Excel 'Code' of 'Lidcode' gebruikt.
    """
    try:
        import pandas as pd
    except Exception as e:
        raise ImportError("pandas is nodig voor load_members_df()") from e

    from config import EXCEL_PATH, SHEET, CODE_COL, NAME_COL

    path = Path(xlsx_path or EXCEL_PATH or (Path(__file__).resolve().parent / "data" / "leden.xlsx"))
    if not path.exists():
        print(f"❌ leden.xlsx niet gevonden: {path}")
        return None

    df = pd.read_excel(path, sheet_name=SHEET, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    cols = list(df.columns)
    norm = {re.sub(r"\s+", "", c).lower(): c for c in cols}

    def pick(wanted: str, fallbacks: list[str]) -> str | None:
        w = re.sub(r"\s+", "", wanted).lower()
        if w in norm:
            return norm[w]
        for f in fallbacks:
            k = re.sub(r"\s+", "", f).lower()
            if k in norm:
                return norm[k]
        return None

    code_src = pick(CODE_COL, ["Lidcode", "Code", "Nummer"])
    name_src = pick(NAME_COL, ["Naam", "Name"])

    if not code_src or not name_src:
        print("❌ Verkeerde kolomnamen in leden.xlsx.")
        print("   Gevonden kolommen:", cols)
        print("   Verwacht minimaal: Code/Lidcode + Naam")
        return None

    if code_src != CODE_COL:
        df.rename(columns={code_src: CODE_COL}, inplace=True)
    if name_src != NAME_COL:
        df.rename(columns={name_src: NAME_COL}, inplace=True)

    return df


def load_back_dates(xlsx_path: Path) -> list[date] | None:
    wb = openpyxl.load_workbook(Path(xlsx_path), data_only=True)

    sheet_name = _find_sheet(wb, ["Planning"])
    if not sheet_name:
        wb.close()
        return None

    ws = wb[sheet_name]
    dates: list[date] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = _to_date(row[0] if row else None)
        if d:
            dates.append(d)

    wb.close()

    if len(dates) >= 10:
        dates.sort()
        return dates[:10]

=======
# excel_loader.py
from __future__ import annotations

from pathlib import Path
from datetime import date, datetime
import re
import unicodedata

import openpyxl


def _normalize_name(s: str) -> str:
    s = re.sub(r"\s+", " ", str(s).strip())
    if " " not in s:
        s = re.sub(r"(?<!^)(?=[A-Z])", " ", s)
    return s.strip()


def compute_checksum_from12(d12: str) -> int:
    s_odd = sum(int(d12[i]) for i in range(0, 12, 2))
    s_even = sum(int(d12[i]) for i in range(1, 12, 2))
    total = s_odd + 3 * s_even
    return (10 - (total % 10)) % 10


def _to_date(v) -> date | None:
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _norm_colname(x) -> str:
    return re.sub(r"\s+", "", str(x).strip().lower())


def _find_col_index(headers: list, wanted: list[str]) -> int | None:
    norm = [_norm_colname(h) for h in headers]
    wanted_norm = [_norm_colname(w) for w in wanted]
    for w in wanted_norm:
        if w in norm:
            return norm.index(w)
    return None


def _find_sheet(wb, preferred: list[str]) -> str | None:
    names = wb.sheetnames
    norm_map = {n.lower().strip(): n for n in names}
    for p in preferred:
        k = p.lower().strip()
        if k in norm_map:
            return norm_map[k]
    return None


def load_members(xlsx_path: Path) -> list[dict]:
    """
    Leest leden uit Excel en geeft terug:
    [{"code13": "...", "name": "...", "qr": "...|None"}, ...]
    Accepteert kolommen:
      - code: "Lidcode" of "Code" of "Nummer"
      - naam: "Naam" of "Name"
      - qr:   "QR" of "Qr" of "Link" (optioneel)
    """
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    sheet_name = _find_sheet(wb, ["Leden", "Members"])
    if not sheet_name:
        wb.close()
        raise FileNotFoundError("Blad 'Leden' ontbreekt in leden.xlsx")

    ws = wb[sheet_name]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_code = _find_col_index(headers, ["Lidcode", "Code", "Nummer"])
    idx_name = _find_col_index(headers, ["Naam", "Name"])
    idx_qr = _find_col_index(headers, ["QR", "Qr", "Link", "Url", "URL"])

    if idx_code is None or idx_name is None:
        wb.close()
        raise KeyError(
            f"Verkeerde kolomnamen in blad '{sheet_name}'. "
            f"Gevonden: {headers}. Verwacht minimaal Code/Lidcode + Naam."
        )

    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        nummer = row[idx_code] if idx_code < len(row) else None
        naam = row[idx_name] if idx_name < len(row) else None
        qr = row[idx_qr] if (idx_qr is not None and idx_qr < len(row)) else None

        if not nummer or not naam:
            continue

        digits = "".join(ch for ch in str(nummer) if ch.isdigit())
        if len(digits) < 12:
            continue

        d12 = digits[:12]
        code13 = d12 + str(compute_checksum_from12(d12))

        out.append(
            {
                "code13": code13,
                "name": _normalize_name(naam),
                "qr": (str(qr).strip() if qr else None),
            }
        )

    wb.close()

    out.sort(
        key=lambda d: unicodedata.normalize("NFKD", d["name"])
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )
    return out


def load_members_df(xlsx_path: Path | None = None):
    """
    Voor menu_lidkaarten.py:
    geeft een DataFrame terug met kolommen CODE_COL en NAME_COL (uit config.py),
    ongeacht of Excel 'Code' of 'Lidcode' gebruikt.
    """
    try:
        import pandas as pd
    except Exception as e:
        raise ImportError("pandas is nodig voor load_members_df()") from e

    from config import EXCEL_PATH, SHEET, CODE_COL, NAME_COL

    path = Path(xlsx_path or EXCEL_PATH or (Path(__file__).resolve().parent / "data" / "leden.xlsx"))
    if not path.exists():
        print(f"❌ leden.xlsx niet gevonden: {path}")
        return None

    df = pd.read_excel(path, sheet_name=SHEET, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    cols = list(df.columns)
    norm = {re.sub(r"\s+", "", c).lower(): c for c in cols}

    def pick(wanted: str, fallbacks: list[str]) -> str | None:
        w = re.sub(r"\s+", "", wanted).lower()
        if w in norm:
            return norm[w]
        for f in fallbacks:
            k = re.sub(r"\s+", "", f).lower()
            if k in norm:
                return norm[k]
        return None

    code_src = pick(CODE_COL, ["Lidcode", "Code", "Nummer"])
    name_src = pick(NAME_COL, ["Naam", "Name"])

    if not code_src or not name_src:
        print("❌ Verkeerde kolomnamen in leden.xlsx.")
        print("   Gevonden kolommen:", cols)
        print("   Verwacht minimaal: Code/Lidcode + Naam")
        return None

    if code_src != CODE_COL:
        df.rename(columns={code_src: CODE_COL}, inplace=True)
    if name_src != NAME_COL:
        df.rename(columns={name_src: NAME_COL}, inplace=True)

    return df


def load_back_dates(xlsx_path: Path) -> list[date] | None:
    wb = openpyxl.load_workbook(Path(xlsx_path), data_only=True)

    sheet_name = _find_sheet(wb, ["Planning"])
    if not sheet_name:
        wb.close()
        return None

    ws = wb[sheet_name]
    dates: list[date] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = _to_date(row[0] if row else None)
        if d:
            dates.append(d)

    wb.close()

    if len(dates) >= 10:
        dates.sort()
        return dates[:10]

>>>>>>> 14b142486c61fce67c54e7dc87a5c29fdb29e6d5
    return None