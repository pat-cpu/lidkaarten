# excel_loader.py
from __future__ import annotations
from pathlib import Path
from datetime import date, datetime
import re, unicodedata
import openpyxl

def _normalize_name(s: str) -> str:
    s = re.sub(r'\s+', ' ', str(s).strip())
    if ' ' not in s:
        s = re.sub(r'(?<!^)(?=[A-Z])', ' ', s)
    return s.strip()

def compute_checksum_from12(d12: str) -> int:
    s_odd  = sum(int(d12[i]) for i in range(0,12,2))
    s_even = sum(int(d12[i]) for i in range(1,12,2))
    total = s_odd + 3*s_even
    return (10 - (total % 10)) % 10

def load_members(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Leden" not in wb.sheetnames:
        wb.close()
        raise FileNotFoundError("Blad 'Leden' ontbreekt in leden.xlsx")
    ws = wb["Leden"]

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nummer, naam, qr = (row + (None, None, None))[:3]
        if not nummer or not naam:
            continue
        digits = ''.join(ch for ch in str(nummer) if ch.isdigit())
        if len(digits) < 12:
            continue

        d12 = digits[:12]
        code13 = d12 + str(compute_checksum_from12(d12))

        out.append({
            "code13": code13,
            "name": _normalize_name(naam),
            "qr": (str(qr).strip() if qr else None)
        })

    wb.close()
    out.sort(key=lambda d: unicodedata.normalize("NFKD", d["name"]).encode("ascii","ignore").decode("ascii").lower())
    return out


# ---- DATE HELPERS ----

def _to_date(v) -> date | None:
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except:
            pass
    return None


def load_back_dates(xlsx_path: Path) -> list[date] | None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "Planning" not in wb.sheetnames:
        wb.close()
        return None

    ws = wb["Planning"]
    dates = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = _to_date(row[0] if row else None)
        if d:
            dates.append(d)

    wb.close()

    if len(dates) >= 10:
        dates = [d if isinstance(d, date) else d.date() for d in dates]
        dates.sort()
        return dates[:10]

    return None
