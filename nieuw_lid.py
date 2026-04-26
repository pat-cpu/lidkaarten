# -*- coding: utf-8 -*-
from pathlib import Path
import re
import pandas as pd
import openpyxl

# === CONFIG ===
from config import EXCEL_PATH, SHEET, CODE_COL, NAME_COL
from datetime import datetime
PREFIX = str(datetime.now().year)

def checksum_ean13(code12: str) -> str:
    if not re.fullmatch(r"\d{12}", code12):
        raise ValueError(f"EAN13 checksum verwacht 12 cijfers, kreeg: {code12!r}")

    total = 0
    for i, c in enumerate(code12):
        n = int(c)
        total += n * 3 if (i + 1) % 2 == 0 else n
    return str((10 - (total % 10)) % 10)

def genereer_volgende_barcode(df: pd.DataFrame) -> str:
    df.columns = df.columns.astype(str).str.strip()

    if CODE_COL not in df.columns:
        raise KeyError(f"Kolom '{CODE_COL}' niet gevonden in tabblad '{SHEET}'. Gevonden: {list(df.columns)}")

    codes = df[CODE_COL].astype(str).str.strip()

    # Exact EAN13: 13 digits, start with PREFIX
    pat = re.compile(rf"^{re.escape(PREFIX)}\d{{9}}$")  # 4 + 9 = 13
    geldige = codes[codes.apply(lambda x: bool(pat.fullmatch(x)))]

    if geldige.empty:
        volgnummer = 0
    else:
        hoogste = geldige.astype(int).max()
        hoogste_str = f"{hoogste:013d}"
        serial8 = int(hoogste_str[4:12])  # PREFIX(4) + serial(8) + checksum(1)
        volgnummer = serial8 + 1

    code12 = PREFIX + f"{volgnummer:08d}"  # 12 digits body
    return code12 + checksum_ean13(code12)

def schrijf_sheet_leden_veilig(path: Path, df: pd.DataFrame, sheet_name: str = SHEET):
    wb = openpyxl.load_workbook(path)
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        ws_old = wb[sheet_name]
        wb.remove(ws_old)
        ws = wb.create_sheet(sheet_name, idx)
    else:
        ws = wb.create_sheet(sheet_name)

    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

    wb.save(path)
    wb.close()

def main():
    print("=== NIEUW LID TOEVOEGEN ===")
    naam = input("Naam (Voornaam Naam): ").strip()
    if not naam:
        print("❌ Geen naam ingegeven. Stop.")
        return

    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET, dtype=str)
    df.columns = df.columns.astype(str).str.strip()

    # --- dubbele naam voorkomen ---
    naam_norm = " ".join(naam.split()).casefold()  # spaties normaliseren + hoofdletterongevoelig
    bestaande_namen = (
            df[NAME_COL]
            .astype(str)
            .fillna("")
        .map(lambda s: " ".join(s.split()).casefold())
    )

    if (bestaande_namen == naam_norm).any():
        bestaand_code = df.loc[bestaande_namen == naam_norm, CODE_COL].iloc[0]
        print(f"❌ '{naam}' staat al in de lijst (Code: {bestaand_code}). Niet toegevoegd.")
        return




    if NAME_COL in df.columns:
        bestaat = df[NAME_COL].astype(str).str.strip().str.lower() == naam.lower()
    if bestaat.any():
        print("❌ Deze naam staat al in de lijst. Niet toegevoegd.")
        return


    nieuwe_code = genereer_volgende_barcode(df)

    nieuw = {CODE_COL: nieuwe_code, NAME_COL: naam}
    df = pd.concat([df, pd.DataFrame([nieuw])], ignore_index=True)

    schrijf_sheet_leden_veilig(EXCEL_PATH, df, SHEET)


    lid_sinds = nieuwe_code[:4]
    print("\n✔ Nieuw lid toegevoegd:")
    print(f"   Naam: {naam}")
    print(f"   Barcode: {nieuwe_code}")
    print(f"   Lid sinds: {lid_sinds}")
    print(f"✔ {EXCEL_PATH.name} bijgewerkt (andere tabbladen behouden)\n")
    



if __name__ == "__main__":
    main()